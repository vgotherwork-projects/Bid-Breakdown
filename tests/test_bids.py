from datetime import date

import pytest
from fastapi.testclient import TestClient

from app.bid_schemas import EmployeeBidInput
from app.calculator import (
    basic_for_ctc,
    calculate_employee_bid,
    employer_pf,
    gratuity_tenure_years,
    tenure_years,
)
from app.main import app

client = TestClient(app)


def rows_by_key(breakdown):
    return {r.key: r for r in breakdown.rows}


@pytest.mark.parametrize(
    "ctc,expected_basic",
    [
        (150000, 15000),      # below first slab -> clamp to 15,000
        (219999, 15000),
        (350000, 15000),
        (630001, 18030),
        (1000000, 22695),
        (1200000, 27588),
        (1440000, 33000),
        (1920000, 42000),
        (2500000, 42000),     # above top slab -> clamp to 42,000
    ],
)
def test_basic_slabs(ctc, expected_basic):
    assert basic_for_ctc(ctc) == expected_basic


def test_employer_pf_variations():
    # Variation 1: basic >= 15000 uses the 15,000 ceiling on the last 0.5%.
    assert employer_pf(27588) == pytest.approx(27588 * 0.12 + 27588 * 0.005 + 15000 * 0.005)
    # Variation 2: basic < 15000 uses basic on all three.
    assert employer_pf(12000) == pytest.approx(12000 * 0.12 + 12000 * 0.005 + 12000 * 0.005)


def test_new_hire_breakdown_is_exact():
    bid = EmployeeBidInput(ctc=1_200_000, employment_type="new_hire", as_of_date=date(2024, 1, 1))
    b = calculate_employee_bid(bid)
    r = rows_by_key(b)

    assert r["basic"].monthly == 27588
    assert r["hra"].monthly == 13794
    assert r["conveyance"].monthly == 1600
    assert r["medical"].monthly == 1250
    assert r["employer_pf"].monthly == 3523.5
    assert r["special_pay"].monthly == 52244.5

    # Total earning always reconciles to CTC.
    assert r["total_earning"].monthly == 100000
    assert r["total_earning"].annual == 1_200_000

    # New hire => tenure defaults to 1 year; PTO now applies (1 yr -> 15 days).
    assert b.tenure_years == 1
    assert b.gratuity_years == 1
    assert b.pto_days == 15
    assert r["pto"].annual == round((100000 / 31) * 15, 2)
    assert r["gratuity"].annual == round(27588 * 1 * (15 / 26), 2)

    expected_grand_annual = round(1_200_000 + r["pto"].annual + r["gratuity"].annual, 2)
    assert r["grand_total"].annual == expected_grand_annual

    # Per-hour column: annual / annual_hours (default 1880).
    assert b.annual_hours == 1880
    assert r["basic"].hourly == round(331056 / 1880, 2)
    assert r["grand_total"].hourly == round(expected_grand_annual / 1880, 2)


def test_annual_hours_is_configurable():
    bid = EmployeeBidInput(
        ctc=1_200_000, employment_type="new_hire", as_of_date=date(2024, 1, 1), annual_hours=2000
    )
    b = calculate_employee_bid(bid)
    r = rows_by_key(b)
    assert b.annual_hours == 2000
    assert r["grand_total"].hourly == round(r["grand_total"].annual / 2000, 2)
    assert r["basic"].hourly == round(331056 / 2000, 2)


def test_new_hire_tenure_defaults_to_one():
    bid = EmployeeBidInput(ctc=1_200_000, employment_type="new_hire", as_of_date=date(2024, 1, 1))
    b = calculate_employee_bid(bid)
    assert b.tenure_years == 1
    assert b.gratuity_years == 1
    assert b.pto_days == 15  # 1 year tenure -> 15 PTO days
    assert rows_by_key(b)["gratuity"].annual == round(27588 * 1 * (15 / 26), 2)


def test_markup_billing_rate_default_25():
    bid = EmployeeBidInput(ctc=1_200_000, employment_type="new_hire", as_of_date=date(2024, 1, 1))
    b = calculate_employee_bid(bid)
    grand = rows_by_key(b)["grand_total"]
    assert b.markup_pct == 25
    # Billing rate is markup applied to the (unrounded) grand-total hourly rate.
    assert b.billing_rate_per_hour == round((grand.annual / b.annual_hours) * 1.25, 2)


def test_markup_is_configurable():
    bid = EmployeeBidInput(
        ctc=1_200_000, employment_type="new_hire", as_of_date=date(2024, 1, 1), markup_pct=40
    )
    b = calculate_employee_bid(bid)
    assert b.markup_pct == 40
    grand = rows_by_key(b)["grand_total"]
    expected = round((grand.annual / b.annual_hours) * 1.40, 2)
    assert b.billing_rate_per_hour == expected


def test_existing_employee_pto_and_gratuity():
    doj = date(2017, 6, 1)
    as_of = date(2024, 6, 1)  # ~7 years tenure
    bid = EmployeeBidInput(
        ctc=1_200_000, employment_type="existing", date_of_joining=doj, as_of_date=as_of
    )
    b = calculate_employee_bid(bid)
    r = rows_by_key(b)

    tenure = tenure_years(doj, as_of)
    assert tenure > 5
    assert b.pto_days == 20
    assert b.gratuity_years == 7  # exactly 7 completed years, no rounding

    expected_pto = round((100000 / 31) * 20, 2)
    expected_gratuity = round(27588 * 7 * (15 / 26), 2)
    assert r["pto"].annual == expected_pto
    assert r["gratuity"].annual == expected_gratuity

    expected_grand = round(1_200_000 + expected_pto + expected_gratuity, 2)
    assert r["grand_total"].annual == pytest.approx(expected_grand, abs=0.05)


@pytest.mark.parametrize(
    "doj,as_of,expected_years",
    [
        (date(2018, 1, 1), date(2024, 1, 1), 6),    # exactly 6 years
        (date(2018, 1, 1), date(2024, 6, 1), 6),    # 6 years 5 months -> down
        (date(2018, 1, 1), date(2024, 7, 1), 6),    # 6 years 6 months -> down (not > 6m)
        (date(2018, 1, 1), date(2024, 8, 1), 7),    # 6 years 7 months -> up
        (date(2018, 1, 1), date(2024, 12, 1), 7),   # 6 years 11 months -> up
        (date(2024, 3, 1), date(2024, 6, 1), 0),    # < 1 year -> 0
    ],
)
def test_gratuity_tenure_rounding(doj, as_of, expected_years):
    assert gratuity_tenure_years(doj, as_of) == expected_years


def test_gratuity_uses_rounded_tenure():
    # 6 years 7 months -> rounds up to 7 for gratuity.
    bid = EmployeeBidInput(
        ctc=1_200_000,
        employment_type="existing",
        date_of_joining=date(2018, 1, 1),
        as_of_date=date(2024, 8, 1),
    )
    b = calculate_employee_bid(bid)
    r = rows_by_key(b)
    assert b.gratuity_years == 7
    assert r["gratuity"].annual == round(27588 * 7 * (15 / 26), 2)


def test_pto_is_15_days_under_5_years():
    bid = EmployeeBidInput(
        ctc=1_200_000,
        employment_type="existing",
        date_of_joining=date(2022, 1, 1),
        as_of_date=date(2024, 1, 1),
    )
    b = calculate_employee_bid(bid)
    assert b.pto_days == 15


def test_recent_existing_employee_gets_min_one_year_gratuity():
    # Joined ~4 months ago: completed years round down to 0, but gratuity
    # must treat any such joiner as a 1-year joiner (never zero).
    bid = EmployeeBidInput(
        ctc=1_200_000,
        employment_type="existing",
        date_of_joining=date(2023, 9, 1),
        as_of_date=date(2024, 1, 1),
    )
    b = calculate_employee_bid(bid)
    assert gratuity_tenure_years(date(2023, 9, 1), date(2024, 1, 1)) == 0
    assert b.gratuity_years == 1
    assert rows_by_key(b)["gratuity"].annual == round(27588 * 1 * (15 / 26), 2)


def test_endpoint_calculates():
    res = client.post(
        "/bids/calculate",
        json={"ctc": 1200000, "employment_type": "new_hire"},
    )
    assert res.status_code == 200
    body = res.json()
    assert body["basic_monthly"] == 27588
    # New hire grand total = CTC + 1-year PTO + 1-year gratuity provision.
    pto = round((100000 / 31) * 15, 2)
    gratuity = round(27588 * 1 * (15 / 26), 2)
    grand = next(r for r in body["rows"] if r["key"] == "grand_total")
    assert grand["annual"] == round(1_200_000 + pto + gratuity, 2)


def test_export_xlsx():
    res = client.post(
        "/bids/export/xlsx",
        json={"name": "Asha Rao", "ctc": 1200000, "employment_type": "new_hire"},
    )
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    disposition = res.headers["content-disposition"]
    assert "Asha Rao_Bid Breakdown_" in disposition
    assert disposition.rstrip().endswith('.xlsx"')
    assert res.content[:2] == b"PK"  # .xlsx is a zip archive


def test_export_xlsx_matches_template():
    import io

    from openpyxl import load_workbook

    # CTC = 800,000 reproduces the reference workbook's hourly figures.
    res = client.post(
        "/bids/export/xlsx",
        json={
            "name": "Ranjetha Priya Arumugam",
            "ctc": 800000,
            "employment_type": "new_hire",
            "as_of_date": "2024-01-01",
        },
    )
    assert res.status_code == 200
    ws = load_workbook(io.BytesIO(res.content))["Bid Breakdown"]

    assert ws["B2"].value == "New Placement"
    assert ws["B4"].value == "Supplier Name"
    assert ws["C4"].value == "STG Infotech (India) LLP"
    assert ws["C5"].value == "Ranjetha Priya Arumugam"
    assert ws["C6"].value == 1880
    assert ws["C7"].value == "Hourly (\u20b9)"
    assert ws["B8"].value == "Worker Payroll (Basic)"
    assert ws["C8"].value == 115.09  # 18030*12/1880
    assert ws["C9"].value == 57.54  # HRA
    assert ws["C14"].value == 7.98  # Health (medical 1250/mo)
    assert ws["C18"].value == 10.21  # Transport (conveyance 1600/mo)
    assert ws["C24"].value == 448.22  # CTC (grand total)
    assert ws["C26"].value == 0.25 and ws["C26"].number_format == "0%"  # Mark-up
    assert ws["D10"].value == "Onetime"  # gratuity note
    assert ws["B8"].font.name == "Aptos"
    assert ws["K67"].value == "Employee Contribution" and ws["L67"].value == 0.12


def test_export_pdf():
    res = client.post(
        "/bids/export/pdf",
        json={"name": "Asha Rao", "ctc": 1200000, "employment_type": "new_hire"},
    )
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/pdf"
    assert "Asha Rao_Bid Breakdown_" in res.headers["content-disposition"]
    assert res.content[:5] == b"%PDF-"


def test_endpoint_existing_requires_doj():
    res = client.post(
        "/bids/calculate",
        json={"ctc": 1200000, "employment_type": "existing"},
    )
    assert res.status_code == 422
