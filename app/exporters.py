"""Excel (.xlsx) and PDF renderers for an employee bid breakdown."""
import io
from functools import lru_cache
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageChops

from .bid_schemas import BatchRowResult, EmployeeBidBreakdown, EmploymentType

LOGO_PATH = Path(__file__).resolve().parent / "static" / "logo.png"

# Pixels whose strongest channel is below this are treated as background.
_DARK_THRESHOLD = 50

# Letterhead / supplier shown on the worker-specific-costs CSV template.
SUPPLIER_NAME = "STG Infotech (India) LLP"

_CURRENCY_SYMBOLS = {"INR": "\u20b9", "USD": "$", "EUR": "\u20ac", "GBP": "\u00a3"}


@lru_cache(maxsize=1)
def _logo_on_white() -> Image.Image | None:
    """Return the logo with its dark background turned white.

    The supplied PNG has a solid black background (no alpha), which fpdf2
    renders as a black box. We map near-black pixels to white so the logo
    blends into the white PDF page, while preserving the blue mark and the
    white "STG" lettering. A pixel counts as background only when *all* of its
    channels are dark, so the strong-blue parts of the mark are kept intact.
    """
    if not LOGO_PATH.exists():
        return None
    logo = Image.open(LOGO_PATH).convert("RGB")
    r, g, b = logo.split()
    max_channel = ImageChops.lighter(ImageChops.lighter(r, g), b)
    background_mask = max_channel.point(lambda p: 255 if p < _DARK_THRESHOLD else 0)
    white = Image.new("RGB", logo.size, (255, 255, 255))
    return Image.composite(white, logo, background_mask)


def _money(value: float, currency: str) -> str:
    return f"{currency} {value:,.2f}"


def _symbol(currency: str) -> str:
    return _CURRENCY_SYMBOLS.get(currency.upper(), currency)


def _country(currency: str) -> str:
    return "India" if currency.upper() == "INR" else currency


# Styling for the .xlsx template (Aptos 10pt). Colours are the workbook's
# theme colours expressed as explicit aRGB: black banner, light-grey labels,
# and Accent-6 (green) value cells in "lighter 60%/80%" tints.
_FONT = "Aptos"
_BLACK = "FF000000"
_WHITE = "FFFFFFFF"
_GRAY = "FFF2F2F2"        # white, 5% darker (theme 0, tint -0.05)
_GREEN6 = "FFC6E0B4"      # accent 6, lighter 60%
_GREEN8 = "FFE2EFDA"      # accent 6, lighter 80%
_THIN = Side(style="thin", color=_BLACK)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hourly_of(b: EmployeeBidBreakdown, key: str):
    for row in b.rows:
        if row.key == key:
            return round(row.hourly, 2)
    return 0


def _write_breakdown_sheet(ws, b: EmployeeBidBreakdown) -> None:
    """Populate `ws` with the styled STG worker-specific-costs layout.

    Labels sit in column B, the hourly (currency) value in column C, and a
    one-off note in column D. Our calculated components map onto the template's
    fixed allowance list; anything we do not model is emitted as 0.
    """
    ws.column_dimensions["B"].width = 31.0
    ws.column_dimensions["C"].width = 21.63

    def put(coord, value, *, fill=None, bold=False, italic=False,
            center=False, border=True, numfmt=None, color=_BLACK):
        cell = ws[coord]
        cell.value = value
        cell.font = Font(name=_FONT, size=10, bold=bold, italic=italic, color=color)
        if fill:
            cell.fill = PatternFill(fill_type="solid", fgColor=fill)
        if border:
            cell.border = _BORDER
        if center:
            cell.alignment = Alignment(horizontal="center")
        if numfmt:
            cell.number_format = numfmt
        return cell

    def hourly(key: str):
        return _hourly_of(b, key)

    placement = (
        "New Placement"
        if b.employment_type == EmploymentType.new_hire
        else "Existing Placement"
    )
    hours = int(b.annual_hours) if float(b.annual_hours).is_integer() else b.annual_hours

    # Banner + meta block (rows 2-7). Style both merged cells, then merge.
    put("B2", placement, fill=_BLACK, bold=True, center=True, color=_WHITE)
    put("C2", None, fill=_BLACK)
    ws.merge_cells("B2:C2")
    put("B3", _country(b.currency), fill=_GRAY, center=True)
    put("C3", None, fill=_GRAY)
    ws.merge_cells("B3:C3")
    put("B4", "Supplier Name", fill=_GRAY)
    put("C4", SUPPLIER_NAME, fill=_GREEN6, bold=True, center=True)
    put("B5", "Worker Name", fill=_GRAY)
    put("C5", b.name, fill=_GREEN6, bold=True, center=True)
    put("B6", "Max Annual Hours", fill=_GRAY)
    put("C6", hours, fill=_GRAY, bold=True, center=True)
    put("B7", "Worker Specific Costs*", fill=_GRAY)
    put("C7", f"Hourly ({_symbol(b.currency)})", fill=_GRAY, bold=True, center=True)

    # Component lines: (row, label, value, value-fill, one-off note).
    lines = [
        (8, "Worker Payroll (Basic)", hourly("basic"), _GREEN6, None),
        (9, "House Rent Allowance (HRA)", hourly("hra"), _GREEN6, None),
        (10, "Gratuity", hourly("gratuity"), _GREEN8, "Onetime"),
        (11, "Provident Fund (PF) - Employers Cont.", hourly("employer_pf"), _GREEN6, None),
        (12, "Bonus", 0, _GREEN8, None),
        (13, "Paid Time Off", hourly("pto"), _GREEN8, "Onetime"),
        (14, "Health Insurance & Life Insurance", hourly("medical"), _GREEN8, None),
        (15, "Driver Allowance", 0, _GREEN8, None),
        (16, "Stationary Allowance", 0, _GREEN8, None),
        (17, "Meal Allowance / Coupons", 0, _GREEN8, None),
        (18, "Transport Allowance", hourly("conveyance"), _GREEN8, None),
        (19, "Internet Allowance", 0, _GREEN8, None),
        (20, "Phone Allowance", 0, _GREEN8, None),
        (21, "Vehicle / Fuel Allowance", 0, _GREEN8, None),
        (22, "Other Worker Specific Cost 1", hourly("special_pay"), _GREEN8, None),
        (23, "Other Worker Specific Cost 2", 0, _GREEN8, None),
    ]
    for r, label, value, fill, note in lines:
        put(f"B{r}", label)
        put(f"C{r}", value, fill=fill)
        if note:
            put(f"D{r}", note, italic=True, border=False)

    put("B24", "CTC")
    put("C24", hourly("grand_total"), fill=_GRAY, bold=True)
    put("B25", "Customer Charge Rate (bid rate)")
    put("C25", round(b.billing_rate_per_hour, 2), fill=_GREEN6)
    put("B26", "Mark-up")
    put("C26", round(b.markup_pct / 100, 4), fill=_GRAY, bold=True, numfmt="0%")

    # Reference rate constants, unstyled, as in the source workbook (K67:L71).
    plain = Font(name=_FONT, size=10)
    for coord, value in (
        ("K67", "Employee Contribution"), ("L67", 0.12),
        ("K68", "Employer Contribution"), ("L68", 0.0833),
        ("L69", 0.0367), ("L70", 0.005), ("L71", 0.005),
    ):
        ws[coord] = value
        ws[coord].font = plain


def breakdown_to_xlsx(b: EmployeeBidBreakdown) -> bytes:
    """Render a single breakdown as a styled .xlsx mirroring the STG workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bid Breakdown"
    _write_breakdown_sheet(ws, b)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Wide master-table columns, in the exact order requested for the batch output.
# Each entry: (header, kind, key, number-format, column-width).
#   kind "id"    -> text attribute carried through from the uploaded file (or blank)
#   kind "num"   -> numeric attribute carried through from the uploaded file (or blank)
#   kind "name"  -> the worker name from the breakdown
#   kind "money" -> the hourly value of a breakdown component (key = row key)
#   kind "zero"  -> a component we do not model; always 0
_BATCH_COLS = [
    ("SL No", "id", "sno", "0", 7),
    ("STGI-ID", "id", "stgi_id", None, 12),
    ("Agency worker Name", "name", None, None, 24),
    ("Supplier", "id", "supplier", None, 22),
    ("B2B Contractor ID", "id", "b2b_id", None, 16),
    ("PO Number", "id", "po_number", None, 14),
    ("PO Rate (INR)", "num", "po_rate", "#,##0.00", 13),
    ("Worker Payroll (Basic)", "money", "basic", "0.00", 14),
    ("House Rent Allowance (HRA)", "money", "hra", "0.00", 16),
    ("Gratuity", "money", "gratuity", "0.00", 11),
    ("Provident Fund (PF) - Employers Cont.", "money", "employer_pf", "0.00", 18),
    ("Bonus", "zero", None, "0.00", 10),
    ("Paid Time Off", "money", "pto", "0.00", 12),
    ("Health Insurance & Life Insurance", "money", "medical", "0.00", 18),
    ("Driver Allowance", "zero", None, "0.00", 13),
    ("Stationary Allowance", "zero", None, "0.00", 14),
    ("Meal Allowance / Coupons", "zero", None, "0.00", 16),
    ("Transport Allowance", "money", "conveyance", "0.00", 14),
    ("Internet Allowance", "zero", None, "0.00", 13),
    ("Phone Allowance", "zero", None, "0.00", 13),
    ("Vehicle / Fuel Allowance", "zero", None, "0.00", 14),
    ("Other Worker Specific Cost 1", "money", "special_pay", "0.00", 16),
    ("Other Worker Specific Cost 2", "zero", None, "0.00", 16),
    ("Supplier Contact Name", "id", "contact_name", None, 20),
    ("Supplier Contact Email", "id", "contact_email", None, 24),
    ("Supplier Contact Phone", "id", "contact_phone", None, 18),
]


def _batch_value(row: BatchRowResult, kind: str, key):
    if kind == "name":
        return row.breakdown.name
    if kind == "id":
        return getattr(row, key) or ""
    if kind == "num":
        value = getattr(row, key)
        return "" if value is None else value
    if kind == "money":
        return _hourly_of(row.breakdown, key)
    return 0  # "zero": an allowance we do not model


def batch_to_xlsx(rows: list[BatchRowResult]) -> bytes:
    """Build one wide master sheet: a row per worker, components across columns."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bid Breakdown"
    ws.freeze_panes = "C2"  # keep SL No / STGI-ID / Name visible while scrolling

    header_font = Font(name=_FONT, size=10, bold=True, color=_WHITE)
    body_font = Font(name=_FONT, size=10)
    header_fill = PatternFill(fill_type="solid", fgColor=_BLACK)
    value_fill = PatternFill(fill_type="solid", fgColor=_GREEN8)
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (header, *_rest) in enumerate(_BATCH_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = _BORDER
        cell.alignment = center_wrap

    for r, row in enumerate(rows, start=2):
        for col_idx, (_header, kind, key, fmt, _w) in enumerate(_BATCH_COLS, start=1):
            cell = ws.cell(row=r, column=col_idx, value=_batch_value(row, kind, key))
            cell.font = body_font
            cell.border = _BORDER
            if fmt:
                cell.number_format = fmt
            if kind in ("money", "zero"):
                cell.fill = value_fill

    for col_idx, (_h, _k, _key, _fmt, width) in enumerate(_BATCH_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 42

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe(text: str) -> str:
    """fpdf2 core fonts are latin-1; replace anything outside it."""
    return text.encode("latin-1", "replace").decode("latin-1")


def breakdown_to_pdf(b: EmployeeBidBreakdown) -> bytes:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Letterhead: logo on the left, title block to its right.
    text_x = pdf.l_margin
    logo = _logo_on_white()
    if logo is not None:
        logo_h = 16.0
        pdf.image(logo, x=pdf.l_margin, y=pdf.get_y(), h=logo_h)
        text_x = pdf.l_margin + 18

    top_y = pdf.get_y()
    pdf.set_xy(text_x, top_y)
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 9, _safe("Employee Cost & Bid Breakdown"), new_x="LMARGIN", new_y="NEXT")
    pdf.set_x(text_x)
    pdf.set_font("Helvetica", "", 11)
    pdf.cell(0, 7, _safe(f"{b.name}  -  {b.currency} {b.ctc:,.0f} CTC ({b.employment_type.value})"),
             new_x="LMARGIN", new_y="NEXT")

    # Ensure content starts below the logo regardless of text height.
    pdf.set_y(max(pdf.get_y(), top_y + 18))
    pdf.set_draw_color(11, 31, 58)
    pdf.set_line_width(0.4)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.ln(3)

    # Meta block
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(90, 90, 90)
    meta = (
        f"DOJ: {b.effective_date_of_joining.isoformat()}   "
        f"Tenure: {b.tenure_years} yrs   Gratuity yrs: {b.gratuity_years}   "
        f"PTO: {b.pto_days} days   Annual hours: {b.annual_hours}   Markup: {b.markup_pct}%"
    )
    pdf.multi_cell(0, 5, _safe(meta))
    pdf.set_text_color(0, 0, 0)
    pdf.ln(3)

    # Table header
    widths = [80, 37, 37, 30]
    headers = ["Component", "Monthly", "Annual", "Per hour"]
    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(11, 31, 58)
    pdf.set_text_color(255, 255, 255)
    for w, h in zip(widths, headers):
        align = "L" if h == "Component" else "R"
        pdf.cell(w, 8, _safe(h), border=0, align=align, fill=True)
    pdf.ln(8)
    pdf.set_text_color(0, 0, 0)

    for row in b.rows:
        is_total = row.kind.value == "total"
        is_sub = row.kind.value == "subtotal"
        pdf.set_font("Helvetica", "B" if (is_total or is_sub) else "", 10)
        if is_total:
            pdf.set_fill_color(238, 244, 255)
            fill = True
        elif is_sub:
            pdf.set_fill_color(245, 247, 251)
            fill = True
        else:
            fill = False
        cells = [
            (widths[0], row.label, "L"),
            (widths[1], _money(row.monthly, b.currency), "R"),
            (widths[2], _money(row.annual, b.currency), "R"),
            (widths[3], _money(row.hourly, b.currency), "R"),
        ]
        for w, text, align in cells:
            pdf.cell(w, 7, _safe(text), border="B", align=align, fill=fill)
        pdf.ln(7)

    pdf.ln(5)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 6, _safe(f"Grand total per hour: {_money(b.grand_total_hourly, b.currency)}"),
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(37, 99, 235)
    pdf.cell(0, 8, _safe(f"Billing rate per hour (+{b.markup_pct}% markup): "
                         f"{_money(b.billing_rate_per_hour, b.currency)}"),
             new_x="LMARGIN", new_y="NEXT")

    out = pdf.output()
    return bytes(out)
