"""Parse an uploaded batch file and compute a breakdown for each worker.

The file is expected to carry a header row. Columns are matched by header name
(case/spacing/punctuation insensitive), so their order is flexible:

    Required to compute a breakdown:
        Name, Date of Joining (yyyy-mm-dd), CTC (annual)
    Optional identity / supplier columns (carried through to the output as-is):
        S. No., STGI-ID, Supplier, B2B Contractor ID, PO Number,
        Supplier Contact Name, Supplier Contact Email, Supplier Contact Phone

If no recognisable header is present we fall back to the legacy positional
layout (S. No., Name, Date of Joining, CTC). Every row is treated as an
existing employee, since a date of joining is given.
"""
import csv
import io
import re
from datetime import date, datetime

from openpyxl import load_workbook

from .bid_schemas import (
    BatchError,
    BatchResult,
    BatchRowResult,
    EmployeeBidInput,
    EmploymentType,
)
from .calculator import calculate_employee_bid

_DATE_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%m/%d/%Y")

# Header aliases, stored already normalised (lowercase, alnum tokens, single spaces).
_HEADER_ALIASES: dict[str, set[str]] = {
    "sno": {"s no", "sno", "sl no", "serial", "serial no", "serial number"},
    "name": {"name", "agency worker name", "worker name", "employee name"},
    "doj": {"date of joining", "doj", "joining date", "date of join"},
    "ctc": {"ctc", "annual ctc", "cost to company"},
    "stgi_id": {"stgi id", "stg id", "stgid"},
    "supplier": {"supplier", "supplier name"},
    "b2b_id": {"b2b contractor id", "b2b contractor cds id", "b2b id", "b2b contractor"},
    "po_number": {"po number", "po", "po no", "purchase order"},
    "po_rate": {"po rate inr", "po rate", "po rate in inr"},
    "contact_name": {"supplier contact name", "contact name"},
    "contact_email": {"supplier contact email", "contact email", "email"},
    "contact_phone": {"supplier contact phone", "contact phone", "phone", "contact number"},
}

_POSITIONAL = {"sno": 0, "name": 1, "doj": 2, "ctc": 3}


def _normalize(text) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(text).strip().lower()).strip()


def _parse_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip().split(" ")[0]
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_ctc(value) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if cleaned in ("", "-", "."):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _coerce_sno(value):
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, int):
        return value
    text = str(value).strip()
    try:
        f = float(text)
        return int(f) if f.is_integer() else text
    except ValueError:
        return text


def _as_text(value) -> str:
    """Stringify an identity/contact cell, keeping integer-like numbers clean."""
    if value in (None, ""):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _map_header(cols: list) -> dict[str, int] | None:
    """Return {field: column index} if `cols` looks like a header row, else None."""
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(cols):
        if cell in (None, ""):
            continue
        norm = _normalize(cell)
        for field, aliases in _HEADER_ALIASES.items():
            if field not in mapping and norm in aliases:
                mapping[field] = idx
                break
    # Trust it as a header only if the columns needed to compute are named.
    if "ctc" in mapping and ("name" in mapping or "doj" in mapping):
        return mapping
    return None


def _looks_like_header(cols: list) -> bool:
    doj = cols[2] if len(cols) > 2 else None
    ctc = cols[3] if len(cols) > 3 else None
    return _parse_ctc(ctc) is None and _parse_date(doj) is None


def _cell(cols: list, idx):
    if idx is None or idx < 0 or idx >= len(cols):
        return None
    value = cols[idx]
    return None if value in (None, "") else value


def _get(cols: list, field: str, header_map: dict[str, int] | None):
    idx = header_map.get(field) if header_map else _POSITIONAL.get(field)
    return _cell(cols, idx)


def _rows_from_bytes(filename: str, data: bytes) -> list[list]:
    name = (filename or "").lower()
    if name.endswith((".csv", ".txt")):
        text = data.decode("utf-8-sig", errors="replace")
        return [list(r) for r in csv.reader(io.StringIO(text))]
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


def process_batch(filename: str, data: bytes) -> BatchResult:
    raw_rows = _rows_from_bytes(filename, data)
    rows = [r for r in raw_rows if any(c not in (None, "") for c in r)]

    header_map: dict[str, int] | None = None
    if rows:
        header_map = _map_header(rows[0])
        if header_map is not None or _looks_like_header(rows[0]):
            rows = rows[1:]

    results: list[BatchRowResult] = []
    errors: list[BatchError] = []

    for i, cols in enumerate(rows, start=1):
        sno = _coerce_sno(_get(cols, "sno", header_map))
        name_val = _get(cols, "name", header_map)
        name = str(name_val).strip() if name_val is not None else ""
        doj = _parse_date(_get(cols, "doj", header_map))
        ctc = _parse_ctc(_get(cols, "ctc", header_map))

        if not name:
            errors.append(BatchError(row=i, sno=sno, message="Missing name"))
            continue
        if doj is None:
            errors.append(BatchError(
                row=i, sno=sno, name=name,
                message="Invalid or missing Date of Joining (expected yyyy-mm-dd)",
            ))
            continue
        if ctc is None or ctc <= 0:
            errors.append(BatchError(row=i, sno=sno, name=name, message="Invalid or missing CTC"))
            continue

        try:
            bid = EmployeeBidInput(
                name=name,
                ctc=ctc,
                employment_type=EmploymentType.existing,
                date_of_joining=doj,
            )
            breakdown = calculate_employee_bid(bid)
        except Exception as exc:  # noqa: BLE001 - surface row-level failures to the user
            errors.append(BatchError(row=i, sno=sno, name=name, message=str(exc)))
            continue

        results.append(BatchRowResult(
            sno=sno,
            stgi_id=_as_text(_get(cols, "stgi_id", header_map)),
            supplier=_as_text(_get(cols, "supplier", header_map)),
            b2b_id=_as_text(_get(cols, "b2b_id", header_map)),
            po_number=_as_text(_get(cols, "po_number", header_map)),
            po_rate=_parse_ctc(_get(cols, "po_rate", header_map)),
            contact_name=_as_text(_get(cols, "contact_name", header_map)),
            contact_email=_as_text(_get(cols, "contact_email", header_map)),
            contact_phone=_as_text(_get(cols, "contact_phone", header_map)),
            breakdown=breakdown,
        ))

    return BatchResult(count=len(results), results=results, errors=errors)
